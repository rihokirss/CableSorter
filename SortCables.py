import pandas as pd
import tkinter as tk
from collections import defaultdict
from tkinter import Tk, filedialog, Toplevel, Listbox, messagebox, END
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Funktsioon kaablite suuna kontrollimiseks ja korrigeerimiseks
def correct_direction(row, direction_counter):
    # Kontrollib, kas "To" sagedus on suurem kui "From" sagedus
    return direction_counter[row['To']] > direction_counter[row['From']]

def sort_cables(input_file_path, sheet_name, output_file_path):
    df = pd.read_excel(input_file_path, sheet_name=sheet_name, engine='openpyxl')
    
    direction_counter = defaultdict(int)
    for _, row in df.iterrows():
        direction_counter[row['From']] += 1
        direction_counter[row['To']] += 1

    for index, row in df.iterrows():
        if correct_direction(row, direction_counter):
            df.at[index, 'From'], df.at[index, 'To'] = row['To'], row['From']

    graph = defaultdict(list)
    edges = set()
    for _, row in df.iterrows():
        graph[row['From']].append(row['To'])
        edges.add((row['From'], row['To']))

    roots = set(graph.keys()) - {to for _, to in edges}

    visited = set()
    sorted_cables = []

    def dfs(node, path=[]):
        if node not in visited:
            visited.add(node)
            path.append(node)
            for next_node in graph[node]:
                dfs(next_node, path)
            if not graph[node]:
                sorted_cables.append(path.copy())
            path.pop()
            visited.remove(node)

    for root in roots:
        dfs(root, [])

    sorted_cables_df = pd.DataFrame(columns=df.columns)
    for path in sorted_cables:
        for i in range(len(path) - 1):
            # Võtke esimene vastav rida
            row = df[(df['From'] == path[i]) & (df['To'] == path[i+1])].iloc[0]
 
            # Eeldame, et sorted_cables_df on teie olemasolev DataFrame ja row on uus rida, mida soovite lisada
            row_df = pd.DataFrame([row])  # Muudab rea DataFrame'iks

            # Eemalda tühjad või NA veerud uuest DataFrame'ist
            row_df.dropna(axis=1, how='all', inplace=True)

            # Seejärel lisage see olemasolevale DataFrame'ile
            sorted_cables_df = pd.concat([sorted_cables_df, row_df], ignore_index=True)


            # Eemaldab duplikaadid, võttes arvesse algust ja lõppu
            #sorted_cables_df = sorted_cables_df.drop_duplicates(subset=['From', 'To'], keep='first')
            # Eemaldab duplikaadid, võttes arvesse ainult ID veergu
            sorted_cables_df = sorted_cables_df.drop_duplicates(subset=['ID'], keep='first')


    book = load_workbook(input_file_path)
    sheet = book[sheet_name]
    column_widths = [max(len(str(cell.value)) for cell in col) for col in sheet.columns]

    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        sorted_cables_df.to_excel(writer, sheet_name=sheet_name, index=False)
        for i, width in enumerate(column_widths, start=1):
            writer.sheets[sheet_name].column_dimensions[get_column_letter(i)].width = width

# Funktsioon töölehe valimiseks dialoogiaknast
def choose_sheet(root, sheet_names):
    def on_select(evt):
        # Valib töölehe, kui kasutaja teeb valiku
        w = evt.widget
        index = int(w.curselection()[0])
        sheet_name = w.get(index)
        root.sheet_name = sheet_name
        top.destroy()

    top = Toplevel(root)
    listbox = Listbox(top)
    listbox.pack(fill="both", expand=True)

    for name in sheet_names:
        listbox.insert(END, name)

    listbox.bind('<<ListboxSelect>>', on_select)
    root.wait_window(top)
    return root.sheet_name

root = Tk()
root.withdraw()  # Peidab Tkinteri põhiakna

# Kasutaja valib sisend- ja väljundfaili
input_file_path = filedialog.askopenfilename(title="Vali sisendfail", filetypes=[("Excel files", "*.xlsx *.xls")])
if input_file_path:
    book = load_workbook(input_file_path, read_only=True)
    sheet_names = book.sheetnames
    root.sheet_name = None  # Lisab atribuudi dünaamiliselt
    sheet_name = choose_sheet(root, sheet_names)

    output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

    # Käivitab sorteerimise, kui on valitud sisendfail, tööleht ja väljundfail
    if sheet_name and output_file_path:
        sort_cables(input_file_path, sheet_name, output_file_path)

##kontrollid
     
# Lae andmed
df_input = pd.read_excel(input_file_path, sheet_name=sheet_name)
df_output = pd.read_excel(output_file_path, sheet_name=sheet_name)

# Kaablite kogus
input_count = len(df_input)
output_count = len(df_output)

# Kasuta 'ID' veergu
id_column = 'ID'

# 2. Kontrolli, kas mõlemas failis on samad kaabli ID-d
input_ids = set(df_input[id_column])
output_ids = set(df_output[id_column])

# Erinevused
missing_in_output = input_ids.difference(output_ids)
missing_in_input = output_ids.difference(input_ids)

# 3. Kontrolli, kas eksisteerib kaableid ilma alguse ja lõputa
# Peame muutma indekseerimise viisi, kasutades listi asemel set'i
missing_start_or_end_in_input = df_input[df_input['From'].isna() | df_input['To'].isna()][id_column].tolist()
missing_start_or_end_in_output = df_output[df_output['From'].isna() | df_output['To'].isna()][id_column].tolist()

# Väljasta tulemused dialoogiaknas
root = tk.Tk()
root.withdraw() # Peidab tkinteri peamise akna

message = (
    f"Kaablite kogus algfailis: {input_count}\n"
    f"Kaablite kogus väljundfailis: {output_count}\n\n"
    f"Kaabli ID-d, mis on algfailis, kuid mitte väljundfailis: {missing_in_output}\n"
    f"Kaabli ID-d, mis on väljundfailis, kuid mitte algfailis: {missing_in_input}\n\n"
    f"Kaablite ID-d ilma alguse ja lõputa algfailis: {missing_start_or_end_in_input}\n"
    f"Kaablite ID-d ilma alguse ja lõputa väljundfailis: {missing_start_or_end_in_output}"
)

messagebox.showinfo("Kontrolli Tulemused", message)

